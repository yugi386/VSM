import cv2
import numpy as np
from pdf2image import convert_from_path, pdfinfo_from_path
import pandas as pd
import matplotlib.pyplot as plt
import os
import subprocess
import sys
import glob
import pytesseract
from openpyxl.styles import Alignment, Font, PatternFill
import gc
import math
import re
from difflib import SequenceMatcher
from concurrent.futures import ThreadPoolExecutor

# === CONFIGURAÇÕES ===
SCRIPT_BASH = "./preparar_batch.sh"
FPS_CONFIG = 1  
MAX_WORKERS = 8

# CONFIGURAÇÃO DE LOG
GERAR_LOG_DEBUG = False  # Mude para True se quiser debug
LOG_FILE = "debug_log.txt"


def log_debug(mensagem):
    """Escreve no arquivo de log se a constante estiver ativa"""
    if not GERAR_LOG_DEBUG:
        return
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(mensagem + "\n")


def executar_batch_bash():
    print(f"--- Iniciando Processamento em Lote (Bash) ---")
    if GERAR_LOG_DEBUG and os.path.exists(LOG_FILE):
        os.remove(LOG_FILE)
    try:
        subprocess.run([SCRIPT_BASH], check=True)
    except Exception as e:
        print(f"Nota: Script bash pulado ou erro: {e}. Continuando...")


def similaridade(a, b):
    return SequenceMatcher(None, a, b).ratio()


def limpar_texto_ocr(texto):
    if not texto:
        return ""
    # Mantém números, letras, acentos e pontuação básica de valores (,.%)
    texto = re.sub(
        r"[^a-zA-Z0-9\sÁÉÍÓÚáéíóúÂÊÔâêôÃÕãõÇç\-\_\.\,\@\:\;\%\$]", "", texto
    )
    return texto.strip()


def ocr_crop(img_crop, config_psm="--psm 7"):
    try:
        gray = cv2.cvtColor(img_crop, cv2.COLOR_BGR2GRAY)
        # Zoom 4x para detalhes
        gray = cv2.resize(
            gray, None, fx=4, fy=4, interpolation=cv2.INTER_CUBIC
        )
        _, thresh = cv2.threshold(
            gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU
        )

        texto = pytesseract.image_to_string(
            thresh, lang="por", config=config_psm
        )
        return limpar_texto_ocr(texto)
    except:
        return ""


# === AQUI ESTAVA O OCR DO QUADRO INTEIRO (obter_texto_completo) ===
# Mantemos a função por compatibilidade, mas NÃO a chamamos mais.
def obter_texto_completo(img_cv):
    """Mantida apenas por compatibilidade – não é mais usada para acelerar o código."""
    try:
        gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
        _, thresh = cv2.threshold(
            gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU
        )
        texto = pytesseract.image_to_string(
            thresh, lang="por", config="--psm 3"
        )
        return limpar_texto_ocr(texto)
    except:
        return ""


def verificar_selecao_texto(img_prev, img_curr, x, y, w, h):
    pad = 2
    x1, y1 = max(0, x - pad), max(0, y - pad)
    x2, y2 = min(img_prev.shape[1], x + w + pad), min(
        img_prev.shape[0], y + h + pad
    )

    roi_prev = img_prev[y1:y2, x1:x2]
    roi_curr = img_curr[y1:y2, x1:x2]

    txt_prev = ocr_crop(roi_prev)
    txt_curr = ocr_crop(roi_curr)

    if len(txt_curr) < 2:
        return False

    sim = similaridade(txt_prev, txt_curr)
    # Texto igual + Mudança Visual = Highlight
    if sim > 0.80:
        return True
    return False


def detectar_texto_crop(img_curr, x, y, w, h):
    pad_x = 60 if w < 30 else 15
    pad_y = 10

    h_img, w_img = img_curr.shape[:2]
    x1 = max(0, x - pad_x)
    y1 = max(0, y - pad_y)
    x2 = min(w_img, x + w + pad_x)
    y2 = min(h_img, y + h + pad_y)

    roi = img_curr[y1:y2, x1:x2]
    texto = ocr_crop(roi)
    return len(texto) >= 1


# === NOVA FUNÇÃO PARA UNIFICAR RODAPÉS PARECIDOS ===
def normalizar_rodape(novo_rodape, lista_conhecida):
    if not novo_rodape or len(novo_rodape) < 3:
        return ""

    # Limpeza extra para rodapé (remove ícones isolados no início ex: "O ", "OQ ")
    rodape_limpo = re.sub(r"^[A-Z]{1,2}\s+", "", novo_rodape).strip()
    if not rodape_limpo:
        rodape_limpo = novo_rodape  # Fallback se limpou demais

    melhor_match = None
    maior_score = 0

    for rodape_existente in lista_conhecida:
        # Compara com o existente
        score = similaridade(
            novo_rodape.lower(), rodape_existente.lower()
        )

        # Também compara sem os ícones iniciais
        rodape_existente_limpo = re.sub(
            r"^[A-Z]{1,2}\s+", "", rodape_existente
        ).strip()
        score_limpo = similaridade(
            rodape_limpo.lower(), rodape_existente_limpo.lower()
        )

        final_score = max(score, score_limpo)

        if final_score > maior_score:
            maior_score = final_score
            melhor_match = rodape_existente

    # Se for > 70% parecido, usa o que já existe para agrupar no Excel
    if maior_score > 0.70:
        return melhor_match

    return novo_rodape


def normalizar_titulo_tela(novo_titulo, lista_conhecida):
    if not novo_titulo:
        return "Desconhecida", False

    # Para normalização, usamos apenas a PRIMEIRA LINHA (Título Superior)
    # Ignora o rodapé para decidir se a tela mudou
    titulo_principal = novo_titulo.split("\n")[0]

    titulo_limpo = re.sub(
        r"^[\s\W_]*(?:o\s|os\s|I\s|X\s|A\s|Y\s|L\s)+",
        "",
        titulo_principal,
        flags=re.IGNORECASE,
    ).strip()
    titulo_limpo = re.sub(r"[\s\W_]+$", "", titulo_limpo)

    if len(titulo_limpo) < 3:
        return "Desconhecida", False

    melhor_match = None
    maior_score = 0

    for titulo_existente in lista_conhecida:
        # Compara apenas com a primeira linha do título existente também
        titulo_existente_principal = titulo_existente.split("\n")[0]

        score = similaridade(
            titulo_limpo.lower(), titulo_existente_principal.lower()
        )
        if score > maior_score:
            maior_score = score
            melhor_match = titulo_existente

    # Se o título superior for muito parecido, mantém o existente
    if maior_score > 0.80:
        return melhor_match, False
    if maior_score > 0.60:
        if titulo_limpo in melhor_match or melhor_match in titulo_limpo:
            return melhor_match, False

    return novo_titulo, True


def processar_ocr_rodape(img_roi):
    """
    Processamento agressivo para limpar ruído e pegar texto de sistema
    """
    if img_roi.size == 0:
        return ""
    try:
        gray = cv2.cvtColor(img_roi, cv2.COLOR_BGR2GRAY)

        # 1. Inverte se for fundo escuro (comum em SAP)
        if np.mean(gray) < 127:
            gray = cv2.bitwise_not(gray)

        # 2. Upscale (3x)
        img_large = cv2.resize(
            gray, None, fx=3, fy=3, interpolation=cv2.INTER_CUBIC
        )

        # 3. Denoising (Remove granulação que vira lixo no OCR)
        img_clean = cv2.fastNlMeansDenoising(img_large, None, 10, 7, 21)

        # 4. Sharpening (Afia as letras para separar do fundo)
        kernel = np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]])
        img_sharp = cv2.filter2D(img_clean, -1, kernel)

        # 5. Threshold (Binarização simples funciona melhor que Otsu para rodapés limpos)
        _, thresh = cv2.threshold(
            img_sharp, 150, 255, cv2.THRESH_BINARY
        )

        # 6. OCR com PSM 6 (Bloco de texto) para ler várias linhas se houver
        texto = pytesseract.image_to_string(
            thresh, lang="por", config="--psm 6"
        )

        # 7. Busca Reversa: Pega a última linha válida (de baixo para cima)
        linhas = texto.split("\n")
        for linha in reversed(linhas):
            limpo = limpar_texto_ocr(linha)

            # Filtro anti-lixo:
            if len(limpo) > 5 and re.search(r"[aeiouAEIOU]", limpo):
                return limpo

        return ""
    except:
        return ""


def obter_rodape_tela(img_cv):
    """
    Recorta os últimos 200px da tela e aplica OCR focado
    """
    height, width, _ = img_cv.shape

    corte_h = 200
    corte_w = width

    y_start = max(height - corte_h, 0)
    y_end = height
    x_start = 0
    x_end = corte_w

    rodape_crop = img_cv[y_start:y_end, x_start:x_end]

    # Salva imagem para debug se necessário
    if GERAR_LOG_DEBUG and rodape_crop.size > 0:
        # Poderia salvar aqui se necessário
        pass

    return processar_ocr_rodape(rodape_crop)


def obter_titulo_tela_aux(img_cv):
    height, width, _ = img_cv.shape
    corte_h = min(150, height)
    cabecalho = img_cv[0:corte_h, 0:width]
    img_gray = cv2.cvtColor(cabecalho, cv2.COLOR_BGR2GRAY)
    if np.mean(img_gray) < 127:
        img_gray = cv2.bitwise_not(img_gray)
    img_large = cv2.resize(
        img_gray, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC
    )
    _, thresh = cv2.threshold(
        img_large, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU
    )
    try:
        texto_bruto = pytesseract.image_to_string(
            thresh, lang="por", config="--psm 6"
        )
        linhas = texto_bruto.split("\n")
        for linha in linhas:
            limpo = limpar_texto_ocr(linha)
            if len(limpo) > 4:
                return limpo
        return limpar_texto_ocr(texto_bruto)
    except:
        return ""


def obter_titulo_tela(img_cv):
    texto_topo = obter_titulo_tela_aux(img_cv)
    # print("Título", texto_topo)

    texto_rodape = obter_rodape_tela(img_cv)
    # print("Rodapé", texto_rodape)

    return (texto_topo if texto_topo else "Sem Titulo", texto_rodape)


def calcular_centro_massa(thresh_img):
    M = cv2.moments(thresh_img)
    if M["m00"] != 0:
        cx = int(M["m10"] / M["m00"])
        cy = int(M["m01"] / M["m00"])
        return cx, cy
    return None, None


def classificar_mudanca_v25_final(
    pagina_pdf,
    img_prev,
    img_curr,
    intensidade_pixels,
    dist_centro,
    thresh_diff,
    diff_gray,
    titulo_prev,
    titulo_curr,
    texto_full_prev,
    texto_full_curr,
):
    """
    Algoritmo v25: Prioridade para Números e Conteúdo Alfanumérico > 1 char.
    Filtra agressivamente caracteres únicos ambíguos em telas estáticas.
    (texto_full_prev/texto_full_curr não são usados no corpo, mas são mantidos
    para compatibilidade de assinatura.)
    """
    h, w = img_prev.shape[:2]
    total_pixels = h * w
    log_prefix = f"[PAG {pagina_pdf}]"

    # 1. MUDANÇA DE TELA VISUAL (Brusca)
    if intensidade_pixels > (total_pixels * 0.25):
        log_debug(f"{log_prefix} DECISAO: MUDANCA DE TELA (Visual)")
        return "MUDANÇA DE TELA"

    # 2. ANALISE DE TÍTULO
    header_h = int(h * 0.15)
    ratio_header = 0
    if header_h > 0:
        header_diff_pixels = np.sum(thresh_diff[0:header_h, :]) / 255
        ratio_header = header_diff_pixels / (header_h * w)

    t_prev_clean = titulo_prev.lower().strip() if titulo_prev else ""
    t_curr_clean = titulo_curr.lower().strip() if titulo_curr else ""

    score_titulo = 1.0
    if len(t_prev_clean) > 3 and len(t_curr_clean) > 3:
        score_titulo = similaridade(t_prev_clean, t_curr_clean)

    if score_titulo < 0.60 and ratio_header > 0.01:
        return "MUDANÇA DE TELA"
    if score_titulo < 0.85 and ratio_header > 0.05:
        return "MUDANÇA DE TELA"

    gray_prev = cv2.cvtColor(img_prev, cv2.COLOR_BGR2GRAY)
    gray_curr = cv2.cvtColor(img_curr, cv2.COLOR_BGR2GRAY)

    # 3. SCROLL
    scroll_bar_roi = thresh_diff[:, -25:]
    scroll_bar_change = np.sum(scroll_bar_roi) / 255
    is_scroll_bar_moving = scroll_bar_change > (h * 0.1)

    if is_scroll_bar_moving and intensidade_pixels > 5000:
        return "SCROLL"

    patch_h = int(h * 0.4)
    margin_w = int(w * 0.2)
    patch = gray_prev[int(h * 0.2) : int(h * 0.2) + patch_h, margin_w : w - margin_w]
    try:
        res = cv2.matchTemplate(
            gray_curr[:, margin_w : w - margin_w], patch, cv2.TM_CCOEFF_NORMED
        )
        _, max_val, _, max_loc = cv2.minMaxLoc(res)
        if max_val > 0.8:
            deslocamento_y = abs(max_loc[1] - int(h * 0.2))
            if deslocamento_y > 10:
                return "SCROLL"
    except:
        pass

    # 4. ANÁLISE DE CONTORNOS
    contours, _ = cv2.findContours(
        thresh_diff, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
    )
    if not contours:
        return "OCIOSO"

    max_contour = max(contours, key=cv2.contourArea)
    x, y, w_rect, h_rect = cv2.boundingRect(max_contour)
    area_rect = w_rect * h_rect

    # 5. DIGITANDO / SELEÇÃO / MOUSE (Lógica Fina)
    if 5 <= h_rect <= 80 and w_rect < (w * 0.8):

        # Análise de Contraste (Evita foco azul pálido)
        mask_roi = thresh_diff[y : y + h_rect, x : x + w_rect]
        diff_roi = diff_gray[y : y + h_rect, x : x + w_rect]
        mean_contrast = 0
        if np.count_nonzero(mask_roi) > 0:
            mean_contrast = cv2.mean(diff_roi, mask=mask_roi)[0]

        if mean_contrast < 50:
            log_debug(
                f"{log_prefix} DECISAO: MOUSE (Contraste Baixo: {mean_contrast:.1f})"
            )
            return "MOUSE / MOVIMENTO"

        # Recorte e OCR Local
        pad_x = 60 if w_rect < 30 else 15
        pad_y = 10
        x1_crop = max(0, x - pad_x)
        y1_crop = max(0, y - pad_y)
        x2_crop = min(w, x + w_rect + pad_x)
        y2_crop = min(h, y + h_rect + pad_y)

        roi_prev = img_prev[y1_crop:y2_crop, x1_crop:x2_crop]
        roi_curr = img_curr[y1_crop:y2_crop, x1_crop:x2_crop]

        txt_prev_raw = ocr_crop(roi_prev)
        txt_curr_raw = ocr_crop(roi_curr)

        # Verifica Highlight (Seleção)
        if w_rect > 20:
            sim_local = similaridade(txt_prev_raw, txt_curr_raw)
            if sim_local > 0.80 and len(txt_curr_raw) > 2:
                return "SELEÇÃO DE TEXTO"

        # Limpeza para comparação
        clean_p = re.sub(r"[^a-zA-Z0-9]", "", txt_prev_raw)
        clean_c = re.sub(r"[^a-zA-Z0-9]", "", txt_curr_raw)

        # Se não tem texto, é só movimento
        if len(clean_c) == 0 and len(clean_p) == 0:
            return "MOUSE / MOVIMENTO"

        # === A REGRA DE OURO v25 ===
        if clean_p != clean_c:

            # 1. REGRA DO NÚMERO (Prioridade Máxima)
            if any(char.isdigit() for char in clean_c):
                log_debug(
                    f"{log_prefix} DECISAO: DIGITANDO (Contém Números: '{clean_c}')"
                )
                return "DIGITANDO"

            # 2. REGRA DE TAMANHO (Consistência)
            if len(clean_c) >= 2:
                log_debug(
                    f"{log_prefix} DECISAO: DIGITANDO (Texto Longo: '{clean_c}')"
                )
                return "DIGITANDO"

            # 3. FILTRO DE 1 CARACTERE (Anti-Cursor)
            chars_suspeitos = [
                "I",
                "l",
                "i",
                "]",
                "[",
                "!",
                "|",
                "f",
                "t",
                "j",
            ]

            if len(clean_c) == 1:
                if clean_c in chars_suspeitos:
                    log_debug(
                        f"{log_prefix} DECISAO: MOUSE (Ignorado char suspeito: '{clean_c}')"
                    )
                    return "MOUSE / MOVIMENTO"
                else:
                    log_debug(
                        f"{log_prefix} DECISAO: DIGITANDO (Char Único Válido: '{clean_c}')"
                    )
                    return "DIGITANDO"

            # Fallback para similaridade
            diff_len = abs(len(clean_p) - len(clean_c))
            sim_clean = similaridade(clean_p, clean_c)
            if diff_len <= 1 and sim_clean > 0.9:
                return "MOUSE / MOVIMENTO"

            return "DIGITANDO"

        return "MOUSE / MOVIMENTO"

    # 6. SUB-TELA vs POP-UP
    if area_rect > (total_pixels * 0.05):
        toca_esquerda = x <= 5
        toca_direita = (x + w_rect) >= (w - 5)
        toca_baixo = (y + h_rect) >= (h - 5)

        if (toca_esquerda or toca_direita or toca_baixo) and y > header_h:
            return "SUB-TELA"

        if not toca_esquerda and not toca_direita and y > header_h:
            return "POP-UP / SISTEMA"

    if (total_pixels * 0.005) < area_rect <= (total_pixels * 0.05):
        if w_rect > 20 and h_rect > 20:
            return "POP-UP / SISTEMA"

    # 7. MOUSE / NAVEGAÇÃO
    if dist_centro > 10:
        return "MOUSE / MOVIMENTO"

    return "CLIQUE / SELEÇÃO"


def analisar_video_inteligente(pdf_path):
    print(f"\n>> Analisando VSM (OCR Rodapé Normalizado - FAST): {pdf_path}...")
    nome_base = os.path.splitext(os.path.basename(pdf_path))[0]

    try:
        info = pdfinfo_from_path(pdf_path)
        total_frames = info["Pages"]
    except:
        return

    dados_vsm = []
    telas_conhecidas_lista = []
    rodapes_conhecidos_lista = []

    prev_frame_bgr = None
    prev_cx, prev_cy = 0, 0

    # Estado inicial (como no analisador12)
    titulo_atual_normalizado = "Inicio"
    telas_conhecidas_lista.append(titulo_atual_normalizado)

    titulo_anterior_bruto = ""
    texto_full_anterior = ""

    LIMIAR_ATIVIDADE = 500
    BATCH_SIZE = 10

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:

        for start_page in range(1, total_frames + 1, BATCH_SIZE):
            last_page = min(start_page + BATCH_SIZE - 1, total_frames)
            print(f"   ...Lote {start_page}-{last_page}...")

            frames = None
            frames_cv = None
            titulos_futuros = None

            try:
                frames = convert_from_path(
                    pdf_path,
                    fmt="bmp",
                    first_page=start_page,
                    last_page=last_page,
                    dpi=300,
                )
            except Exception as e:
                print(e)
                break

            if not frames:
                continue

            frames_cv = [
                cv2.cvtColor(np.array(f), cv2.COLOR_RGB2BGR)
                for f in frames
            ]

            # OCR em threads – SOMENTE TÍTULO+RODAPÉ
            titulos_futuros = list(executor.map(obter_titulo_tela, frames_cv))

            # >>> NÃO fazemos mais OCR de página inteira aqui <<<
            # textos_futuros = list(executor.map(obter_texto_completo, frames_cv))

            for i, curr_frame_bgr in enumerate(frames_cv):

                frame_index = (start_page - 1) + i
                pagina_pdf = frame_index + 1
                tempo_segundos = frame_index / FPS_CONFIG

                status = "OCIOSO"
                classe = "NVA"
                cx, cy = 0, 0
                intensidade_sum = 0
                dist_movimento = 0

                if titulos_futuros and i < len(titulos_futuros):
                    titulo_bruto, rodape_atual = titulos_futuros[i]
                else:
                    titulo_bruto, rodape_atual = "", ""

                # Texto full agora é sempre vazio (não era usado na lógica)
                texto_full_atual = ""

                # Normalização de título (pelo topo)
                if titulo_bruto and len(titulo_bruto) > 3:
                    nome_normalizado, atualizar = normalizar_titulo_tela(
                        titulo_bruto, telas_conhecidas_lista
                    )
                    if atualizar:
                        if nome_normalizado not in telas_conhecidas_lista:
                            telas_conhecidas_lista.append(nome_normalizado)
                        titulo_atual_normalizado = nome_normalizado
                    else:
                        titulo_atual_normalizado = nome_normalizado

                # Rodapé normalizado
                if rodape_atual and len(rodape_atual) > 3:
                    rodape_norm = normalizar_rodape(
                        rodape_atual, rodapes_conhecidos_lista
                    )
                    if rodape_norm not in rodapes_conhecidos_lista:
                        rodapes_conhecidos_lista.append(rodape_norm)
                    rodape_atual = rodape_norm

                if prev_frame_bgr is not None:
                    gray_prev = cv2.cvtColor(
                        prev_frame_bgr, cv2.COLOR_BGR2GRAY
                    )
                    gray_curr = cv2.cvtColor(
                        curr_frame_bgr, cv2.COLOR_BGR2GRAY
                    )
                    diff = cv2.absdiff(gray_prev, gray_curr)

                    _, thresh = cv2.threshold(
                        diff, 20, 255, cv2.THRESH_BINARY
                    )
                    intensidade_sum = np.sum(thresh) / 255

                    cx, cy = calcular_centro_massa(thresh)
                    if cx is not None and prev_cx is not None:
                        dist_movimento = math.sqrt(
                            (cx - prev_cx) ** 2 + (cy - prev_cy) ** 2
                        )

                    if intensidade_sum > 10:
                        tipo = classificar_mudanca_v25_final(
                            pagina_pdf,
                            prev_frame_bgr,
                            curr_frame_bgr,
                            intensidade_sum,
                            dist_movimento,
                            thresh,
                            diff,
                            titulo_anterior_bruto,
                            titulo_bruto,
                            texto_full_anterior,
                            texto_full_atual,
                        )

                        status = tipo
                        if tipo == "MUDANÇA DE TELA":
                            classe = "Operação (Necessária)"
                        elif tipo == "POP-UP / SISTEMA":
                            classe = "Operação (Necessária)"
                        elif tipo == "SUB-TELA":
                            classe = "Operação (Necessária)"
                        elif tipo == "SCROLL":
                            classe = "Tempo para ser Otimizado"
                        elif tipo == "MOUSE / MOVIMENTO":
                            classe = "Operação (Necessária)"
                        elif tipo == "CLIQUE / SELEÇÃO":
                            classe = "Operação (Necessária)"
                        elif tipo == "SELEÇÃO DE TEXTO":
                            classe = "Operação (Necessária)"
                        elif tipo == "DIGITANDO":
                            classe = "Valor Agregado (VA)"

                # Atualiza estado anterior
                if titulo_bruto and len(titulo_bruto) > 3:
                    titulo_anterior_bruto = titulo_bruto

                texto_full_anterior = texto_full_atual

                prev_cx = cx if cx is not None else prev_cx
                prev_cy = cy if cy is not None else prev_cy
                prev_frame_bgr = curr_frame_bgr

                mins, secs = divmod(int(tempo_segundos), 60)
                timestamp_str = f"{mins:02d}:{secs:02d}"
                intensidade_grafico = int(intensidade_sum)

                dados_vsm.append(
                    {
                        "Pagina_PDF": pagina_pdf,
                        "Tempo_Real_Seg": tempo_segundos,
                        "Timestamp": timestamp_str,
                        "Intensidade": intensidade_grafico,
                        "Status": status,
                        "Classificacao_VSM": classe,
                        "Tela_Nome_Normalizado": titulo_atual_normalizado,
                        "Rodape": rodape_atual,
                    }
                )

            if frames:
                del frames
            if frames_cv:
                del frames_cv
            if titulos_futuros:
                del titulos_futuros
            gc.collect()

    if not dados_vsm:
        return

    # === A PARTIR DAQUI: MESMO CÓDIGO DO analisador12 ORIGINAL ===

    dicionario_telas = {}
    mapa_nome_para_id = {}
    contador_telas = 1
    for nome in telas_conhecidas_lista:
        codigo = f"T{contador_telas:02d}"
        dicionario_telas[codigo] = nome
        mapa_nome_para_id[nome] = codigo
        contador_telas += 1

    for linha in dados_vsm:
        nome = linha["Tela_Nome_Normalizado"]
        linha["Tela_ID"] = mapa_nome_para_id.get(nome, "T00")

    df = pd.DataFrame(dados_vsm)

    # === 1. GRÁFICO ===
    plt.figure(figsize=(18, 10))
    max_int = df["Intensidade"].max() if df["Intensidade"].max() > 0 else 1

    df_dig = df[(df["Status"] == "DIGITANDO")]
    plt.scatter(
        df_dig["Tempo_Real_Seg"],
        df_dig["Intensidade"],
        label="Digitação (VA)",
        color="#1f77b4",
        marker="s",
        s=40,
    )

    df_clique = df[(df["Status"] == "CLIQUE / SELEÇÃO")]
    plt.scatter(
        df_clique["Tempo_Real_Seg"],
        df_clique["Intensidade"],
        label="Clique/Seleção",
        color="blue",
        marker="*",
        s=80,
    )

    df_mouse = df[(df["Status"] == "MOUSE / MOVIMENTO")]
    plt.scatter(
        df_mouse["Tempo_Real_Seg"],
        df_mouse["Intensidade"],
        label="Mouse (Movimento)",
        color="green",
        s=15,
        alpha=0.3,
    )

    df_sel = df[(df["Status"] == "SELEÇÃO DE TEXTO")]
    plt.scatter(
        df_sel["Tempo_Real_Seg"],
        df_sel["Intensidade"],
        label="Seleção Texto",
        color="cyan",
        marker="|",
        s=100,
    )

    mudancas = df[df["Status"] == "MUDANÇA DE TELA"]
    for idx, row in mudancas.iterrows():
        plt.axvline(
            x=row["Tempo_Real_Seg"], color="red", alpha=0.6, linewidth=2
        )
        plt.text(
            row["Tempo_Real_Seg"],
            max_int,
            row["Tela_ID"],
            rotation=90,
            color="red",
            fontsize=9,
            verticalalignment="top",
            fontweight="bold",
        )

    popups = df[df["Status"].isin(["POP-UP / SISTEMA", "SUB-TELA"])]
    for idx, row in popups.iterrows():
        plt.axvline(
            x=row["Tempo_Real_Seg"],
            color="purple",
            alpha=0.6,
            linestyle=":",
        )

    scrolls = df[df["Status"] == "SCROLL"]
    for idx, row in scrolls.iterrows():
        plt.axvline(
            x=row["Tempo_Real_Seg"],
            color="orange",
            alpha=0.4,
            linestyle="--",
        )

    from matplotlib.lines import Line2D

    custom_lines = [
        Line2D([0], [0], color="#1f77b4", marker="s", lw=0),
        Line2D([0], [0], color="blue", marker="*", lw=0, markersize=10),
        Line2D([0], [0], color="green", marker="o", lw=0),
        Line2D([0], [0], color="cyan", marker="|", lw=0, markersize=10),
        Line2D([0], [0], color="red", lw=2),
        Line2D([0], [0], color="purple", linestyle=":", lw=2),
        Line2D([0], [0], color="orange", linestyle="--", lw=2),
    ]
    plt.legend(
        custom_lines,
        [
            "Digitação",
            "Clique/Seleção",
            "Mouse",
            "Seleção Texto",
            "Troca de Tela",
            "Pop-up/Sub-tela",
            "Scroll",
        ],
    )

    plt.title(f"VSM Inteligente v25: {nome_base}", fontsize=14)
    plt.xlabel("Tempo Real (Segundos)")
    plt.ylabel("Intensidade de Mudança (Pixels)")
    plt.subplots_adjust(bottom=0.1, top=0.9, left=0.08, right=0.95)

    tempo_total_seg = df["Tempo_Real_Seg"].max()
    step = 5 if tempo_total_seg > 20 else 1
    plt.xticks(np.arange(0, tempo_total_seg + 5, step))
    plt.grid(True, alpha=0.2)
    plt.savefig(f"grafico_{nome_base}.png", dpi=150)
    plt.close()

    # === 2. EXCEL ===
    arquivo_excel = f"relatorio_{nome_base}.xlsx"
    with pd.ExcelWriter(arquivo_excel, engine="openpyxl") as writer:

        total_tempo_seg = len(df) / FPS_CONFIG

        frames_input = len(df[df["Status"] == "DIGITANDO"])
        frames_operacao = len(
            df[
                df["Status"].isin(
                    [
                        "MUDANÇA DE TELA",
                        "POP-UP / SISTEMA",
                        "SUB-TELA",
                        "CLIQUE / SELEÇÃO",
                        "MOUSE / MOVIMENTO",
                        "SELEÇÃO DE TEXTO",
                    ]
                )
            ]
        )
        frames_otimizar = len(df[df["Status"].isin(["SCROLL"])])
        frames_ocioso = len(df[df["Status"] == "OCIOSO"])

        tempo_input = frames_input / FPS_CONFIG
        tempo_operacao = frames_operacao / FPS_CONFIG
        tempo_otimizar = frames_otimizar / FPS_CONFIG
        tempo_ocioso = frames_ocioso / FPS_CONFIG

        eficiencia = 0
        if total_tempo_seg > 0:
            eficiencia = (
                (tempo_input + tempo_operacao) / total_tempo_seg
            ) * 100

        resumo_dados = {
            "Métrica": [
                "Tempo Total do Processo",
                "Tempo de Input (Digitação - VA)",
                "Tempo de Operação (Navegação/Sistêmico)",
                "Tempo para ser Otimizado (Scroll)",
                "Tempo Ocioso (Leitura/Pausa)",
                "Eficiência Operacional (%)",
            ],
            "Valor": [
                f"{total_tempo_seg:.1f} seg",
                f"{tempo_input:.1f} seg",
                f"{tempo_operacao:.1f} seg",
                f"{tempo_otimizar:.1f} seg",
                f"{tempo_ocioso:.1f} seg",
                f"{eficiencia:.1f}%",
            ],
            "Descrição": [
                "Lead Time total (Frames / FPS)",
                "Tempo efetivo inserindo dados",
                "Tempo navegando, selecionando campos, mouse e processamento",
                "Tempo gasto principalmente com rolagem excessiva",
                "Tempo sem atividade detectada",
                "Percentual de tempo produtivo",
            ],
        }
        pd.DataFrame(resumo_dados).to_excel(
            writer, sheet_name="Resumo de Dados", index=False
        )

        colunas_finais = [
            "Pagina_PDF",
            "Tempo_Real_Seg",
            "Timestamp",
            "Intensidade",
            "Status",
            "Classificacao_VSM",
            "Tela_ID",
        ]
        df[colunas_finais].to_excel(
            writer, sheet_name="Timeline Detalhada", index=False
        )

        catalogo = []
        tela_anterior = None
        inicio_pag = 1
        for idx, row in df.iterrows():
            atual = row["Tela_ID"]
            pag = row["Pagina_PDF"]
            if atual != tela_anterior:
                if tela_anterior is not None:
                    catalogo.append(
                        {
                            "Pagina Inicial": inicio_pag,
                            "Pagina Final": pag - 1,
                            "Tela_ID": tela_anterior,
                        }
                    )
                tela_anterior = atual
                inicio_pag = pag
        catalogo.append(
            {
                "Pagina Inicial": inicio_pag,
                "Pagina Final": len(df),
                "Tela_ID": tela_anterior,
            }
        )
        pd.DataFrame(catalogo).to_excel(
            writer, sheet_name="Catálogo Telas", index=False
        )

        df_dic = pd.DataFrame(
            list(dicionario_telas.items()),
            columns=["Código", "Nome da Tela"],
        )
        df_dic.to_excel(
            writer, sheet_name="Dicionário Telas", index=False
        )

        # === ABA TEMPO PROCESSAMENTO (Com subdivisão de Rodapé) ===
        lista_tempo_telas = []
        if dados_vsm:
            idx_inicio = 0
            nome_tela_atual = dados_vsm[0]["Tela_Nome_Normalizado"]
            rodape_atual = dados_vsm[0]["Rodape"]
            popups_detectados = []

            for i, linha in enumerate(dados_vsm):
                novo_nome = linha["Tela_Nome_Normalizado"]
                novo_rodape = linha["Rodape"]

                if (novo_nome != nome_tela_atual) or (
                    novo_rodape != rodape_atual
                ):

                    slide_ini = dados_vsm[idx_inicio]["Pagina_PDF"]
                    slide_fim = dados_vsm[i - 1]["Pagina_PDF"]
                    tempo = (i - idx_inicio) / FPS_CONFIG
                    str_popups = (
                        ", ".join(
                            map(
                                str,
                                sorted(list(set(popups_detectados))),
                            )
                        )
                        if popups_detectados
                        else "-"
                    )

                    nome_exibicao = nome_tela_atual
                    if rodape_atual:
                        nome_exibicao += f"\n[Rodapé]: {rodape_atual}"

                    lista_tempo_telas.append(
                        {
                            "Nome da Tela + Rodapé": nome_exibicao,
                            "Slide Inicial": slide_ini,
                            "Slide Final": slide_fim,
                            "Pop-ups Detectados (Slides)": str_popups,
                            "Tempo de Processo (s)": tempo,
                        }
                    )

                    idx_inicio = i
                    nome_tela_atual = novo_nome
                    rodape_atual = novo_rodape
                    popups_detectados = []

                if linha["Status"] in ["POP-UP / SISTEMA", "SUB-TELA"]:
                    popups_detectados.append(linha["Pagina_PDF"])

            slide_ini = dados_vsm[idx_inicio]["Pagina_PDF"]
            slide_fim = dados_vsm[-1]["Pagina_PDF"]
            tempo = (len(dados_vsm) - idx_inicio) / FPS_CONFIG
            str_popups = (
                ", ".join(
                    map(str, sorted(list(set(popups_detectados))))
                )
                if popups_detectados
                else "-"
            )

            nome_exibicao = nome_tela_atual
            if rodape_atual:
                nome_exibicao += f"\n[Rodapé]: {rodape_atual}"

            lista_tempo_telas.append(
                {
                    "Nome da Tela + Rodapé": nome_exibicao,
                    "Slide Inicial": slide_ini,
                    "Slide Final": slide_fim,
                    "Pop-ups Detectados (Slides)": str_popups,
                    "Tempo de Processo (s)": tempo,
                }
            )

            pd.DataFrame(lista_tempo_telas).to_excel(
                writer,
                sheet_name="Tempo de processo por tela",
                index=False,
            )

        wb = writer.book
        ws_resumo = wb["Resumo de Dados"]
        ws_resumo.column_dimensions["A"].width = 45
        ws_resumo.column_dimensions["B"].width = 20
        ws_resumo.column_dimensions["C"].width = 80

        ws_tempo = wb["Tempo de processo por tela"]
        ws_tempo.column_dimensions["A"].width = 60
        ws_tempo.column_dimensions["B"].width = 15
        ws_tempo.column_dimensions["C"].width = 15
        ws_tempo.column_dimensions["D"].width = 30
        ws_tempo.column_dimensions["E"].width = 30

        ws = wb["Timeline Detalhada"]
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 30
        ws.column_dimensions["G"].width = 15

        red_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )
        green_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        blue_fill = PatternFill(
            start_color="DAE8FC", end_color="DAE8FC", fill_type="solid"
        )
        purple_fill = PatternFill(
            start_color="E1D5E7", end_color="E1D5E7", fill_type="solid"
        )
        yellow_fill = PatternFill(
            start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
        )

        for row in ws.iter_rows(min_row=2, max_row=len(df) + 1):
            val = row[4].value
            if val == "MUDANÇA DE TELA":
                for cell in row:
                    cell.fill = red_fill
            elif val == "DIGITANDO":
                for cell in row:
                    cell.fill = green_fill
            elif val == "CLIQUE / SELEÇÃO":
                for cell in row:
                    cell.fill = blue_fill
            elif val in ["POP-UP / SISTEMA", "SUB-TELA"]:
                for cell in row:
                    cell.fill = purple_fill
            elif val == "SELEÇÃO DE TEXTO":
                for cell in row:
                    cell.fill = yellow_fill

        ws_dic = wb["Dicionário Telas"]
        ws_dic.column_dimensions["B"].width = 100
        for row in ws_dic.iter_rows(min_row=2):
            row[1].alignment = Alignment(
                wrap_text=True, vertical="center"
            )

        ws_tela = wb["Catálogo Telas"]
        ws_tela.column_dimensions["A"].width = 15
        ws_tela.column_dimensions["B"].width = 15
        ws_tela.column_dimensions["C"].width = 15

        ws_tempo = wb["Tempo de processo por tela"]
        ws_tempo.column_dimensions["A"].width = 100
        for row in ws_tempo.iter_rows(min_row=2):
            row[0].alignment = Alignment(
                wrap_text=True, vertical="center"
            )

    print(f"   -> Sucesso: relatorio_{nome_base}.xlsx")


if __name__ == "__main__":
    executar_batch_bash()
    pdfs = glob.glob("*.pdf")
    for pdf in pdfs:
        analisar_video_inteligente(pdf)

